from __future__ import annotations

import sys
import os
import traceback
import pandas as pd

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QTableView, QFileDialog, QAction,
    QStatusBar, QLabel, QMessageBox, QFrame, QVBoxLayout,
    QHBoxLayout, QLineEdit, QPushButton, QTabBar,
    QColorDialog, QSpinBox, QMenu, QInputDialog, QWidget,
)
from PyQt5.QtGui import (
    QKeySequence, QStandardItemModel, QStandardItem,
    QPainter, QFont, QColor, QBrush, QPixmap, QIcon,
)
from PyQt5.QtCore import Qt, QPoint, QTimer, QSize, QEvent, QRect

from .constants import SUPPORTED_EXTENSIONS, CSV_ENCODINGS, _FMT_KEYS_ROLE
from .utils import load_icon, load_pixmap, _column_is_numeric
from .models import MultiColumnFilterProxyModel
from .filter_popup import FilterPopup
from .overlay import MarchingAntsOverlay
from .header import SortFilterHeaderView


class TableViewerApp(QMainWindow):

    def __init__(self):
        super().__init__()
        self.df                = None
        self.current_file_path = None
        self._source_model     = None
        self._proxy_model      = None
        self._filter_popup     = None   # keep reference – prevents GC
        self._excel_sheets: list      = []
        self._sheet_cache:  dict      = {}
        self._all_sheet_formats: dict = {}
        self._user_changes: dict      = {}
        self._structural_ops: list    = []
        self._undo_stack: list        = []
        self._cell_shadow: dict       = {}
        self._fmt_fg_color: QColor    = QColor(Qt.black)
        self._fmt_bg_color: QColor    = QColor(Qt.white)
        self._clipboard_region        = None
        self._march_offset: float     = 0.0
        self._init_ui()
        self._load_blank_sheet()

    def _init_ui(self):
        self.setGeometry(200, 150, 1100, 680)
        self.setWindowTitle("Table Viewer")
        self.setWindowIcon(load_icon("favicon.ico"))
        self.setAcceptDrops(True)

        self._header = SortFilterHeaderView()
        self._header.sort_requested.connect(self._on_sort_requested)
        self._header.filter_requested.connect(self._show_filter_popup)
        self._header.clear_filter_requested.connect(self._on_clear_filter)

        self.table_view = QTableView(self)
        self.table_view.setHorizontalHeader(self._header)
        self.table_view.horizontalHeader().setSectionsMovable(True)
        self.table_view.setSortingEnabled(False)
        self.table_view.setAlternatingRowColors(True)
        self.table_view.setSelectionBehavior(QTableView.SelectItems)
        self.table_view.setSelectionMode(QTableView.ExtendedSelection)
        self.table_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_view.customContextMenuRequested.connect(self._show_context_menu)
        self.table_view.verticalHeader().setDefaultSectionSize(22)
        self.table_view.verticalHeader().setMinimumWidth(48)
        self.table_view.setViewportMargins(0, 0, 10, 0)

        # Marching-ants overlay
        self._march_overlay = MarchingAntsOverlay(self.table_view.viewport())
        self._march_timer = QTimer(self)
        self._march_timer.setInterval(80)
        self._march_timer.timeout.connect(self._march_tick)
        self.table_view.viewport().installEventFilter(self)
        self.table_view.horizontalScrollBar().valueChanged.connect(self._update_march_overlay)
        self.table_view.verticalScrollBar().valueChanged.connect(self._update_march_overlay)

        # Search bar
        search_bar = QWidget()
        search_bar.setContentsMargins(0, 0, 0, 0)
        sb_layout = QHBoxLayout(search_bar)
        sb_layout.setContentsMargins(8, 7, 8, 7)
        sb_layout.setSpacing(6)

        search_icon_lbl = QLabel()
        search_icon_lbl.setPixmap(load_pixmap("search.ico", 22, 22))
        sb_layout.addWidget(search_icon_lbl)

        self.search_clear_btn = QPushButton()
        self.search_clear_btn.setIcon(load_icon("cancel.ico"))
        self.search_clear_btn.setIconSize(QSize(14, 14))
        self.search_clear_btn.setFixedSize(22, 22)
        self.search_clear_btn.setFlat(True)
        self.search_clear_btn.setVisible(False)
        self.search_clear_btn.setToolTip("Clear search")
        self.search_clear_btn.clicked.connect(self._clear_global_search)
        sb_layout.addWidget(self.search_clear_btn)

        self.global_search_input = QLineEdit()
        self.global_search_input.setPlaceholderText("Search in all columns…")
        self.global_search_input.setMaximumWidth(300)
        self.global_search_input.setMinimumHeight(30)
        self.global_search_input.textChanged.connect(self._on_global_search_text_changed)
        sb_layout.addWidget(self.global_search_input)
        sb_layout.addStretch()

        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(500)
        self._search_timer.timeout.connect(self._apply_global_search)

        # Sheet tab bar
        self.sheet_tab_bar = QTabBar()
        self.sheet_tab_bar.setShape(QTabBar.RoundedSouth)
        self.sheet_tab_bar.setVisible(False)
        self.sheet_tab_bar.currentChanged.connect(self._on_sheet_tab_changed)

        self._format_toolbar = self._build_format_toolbar()

        container = QWidget()
        c_layout = QVBoxLayout(container)
        c_layout.setContentsMargins(0, 0, 0, 0)
        c_layout.setSpacing(0)
        c_layout.addWidget(search_bar)
        c_layout.addWidget(self._format_toolbar)
        c_layout.addWidget(self.table_view)
        c_layout.addWidget(self.sheet_tab_bar)
        self.setCentralWidget(container)

        self.status_label = QLabel("No file loaded")
        status_bar = QStatusBar(self)
        status_bar.addWidget(self.status_label)
        self.setStatusBar(status_bar)

        self._build_menu()

    def _build_menu(self):
        menu_bar = self.menuBar()

        file_menu = menu_bar.addMenu("File")

        open_action = QAction(load_icon("search.ico"), "Open…", self)
        open_action.setShortcut(QKeySequence("Ctrl+O"))
        open_action.triggered.connect(self.show_open_dialog)
        file_menu.addAction(open_action)

        file_menu.addSeparator()

        save_excel = QAction("Save as Excel (.xlsx)", self)
        save_excel.setShortcut(QKeySequence("Ctrl+S"))
        save_excel.triggered.connect(self.save_as_excel)
        file_menu.addAction(save_excel)

        save_csv = QAction("Save as CSV (.csv)", self)
        save_csv.setShortcut(QKeySequence("Ctrl+Shift+S"))
        save_csv.triggered.connect(self.save_as_csv)
        file_menu.addAction(save_csv)

        tools_menu = menu_bar.addMenu("Tools")
        reg_action = QAction(
            load_icon("apps.ico"),
            "Register File Associations (.xlsx .xls .csv)", self
        )
        reg_action.triggered.connect(self.register_file_associations)
        tools_menu.addAction(reg_action)

        undo_action = QAction("Undo", self)
        undo_action.setShortcut(QKeySequence.Undo)
        undo_action.triggered.connect(self._undo)
        self.addAction(undo_action)

        copy_action = QAction("Copy", self)
        copy_action.setShortcut(QKeySequence.Copy)
        copy_action.triggered.connect(self._copy_selection)
        self.addAction(copy_action)

        cut_action = QAction("Cut", self)
        cut_action.setShortcut(QKeySequence.Cut)
        cut_action.triggered.connect(self._cut_selection)
        self.addAction(cut_action)

        paste_action = QAction("Paste", self)
        paste_action.setShortcut(QKeySequence.Paste)
        paste_action.triggered.connect(self._paste_selection)
        self.addAction(paste_action)

        escape_action = QAction("Cancel Copy/Cut", self)
        escape_action.setShortcut(QKeySequence(Qt.Key_Escape))
        escape_action.triggered.connect(self._cancel_copy_cut)
        self.addAction(escape_action)

    # ------------------------------------------------------------------
    # Format toolbar
    # ------------------------------------------------------------------

    def _build_format_toolbar(self) -> QWidget:
        bar = QWidget()
        bar.setContentsMargins(0, 0, 0, 0)
        bar.setMaximumHeight(38)

        layout = QHBoxLayout(bar)
        layout.setContentsMargins(6, 3, 6, 3)
        layout.setSpacing(4)

        self.fmt_bold_btn = QPushButton("B")
        self.fmt_bold_btn.setCheckable(True)
        self.fmt_bold_btn.setFixedSize(26, 26)
        self.fmt_bold_btn.setToolTip("Bold")
        bold_font = QFont(self.fmt_bold_btn.font())
        bold_font.setBold(True)
        self.fmt_bold_btn.setFont(bold_font)
        self.fmt_bold_btn.setStyleSheet(
            "QPushButton{border:1px solid #aaa;border-radius:3px;}"
            "QPushButton:checked{background:#0078d4;color:white;border-color:#005a9e;}"
            "QPushButton:hover:!checked{background:#e0e0e0;}"
        )
        self.fmt_bold_btn.toggled.connect(
            lambda checked: self._apply_format_to_selection('bold', checked)
        )
        layout.addWidget(self.fmt_bold_btn)

        self.fmt_italic_btn = QPushButton("I")
        self.fmt_italic_btn.setCheckable(True)
        self.fmt_italic_btn.setFixedSize(26, 26)
        self.fmt_italic_btn.setToolTip("Italic")
        italic_font = QFont(self.fmt_italic_btn.font())
        italic_font.setItalic(True)
        self.fmt_italic_btn.setFont(italic_font)
        self.fmt_italic_btn.setStyleSheet(
            "QPushButton{border:1px solid #aaa;border-radius:3px;}"
            "QPushButton:checked{background:#0078d4;color:white;border-color:#005a9e;}"
            "QPushButton:hover:!checked{background:#e0e0e0;}"
        )
        self.fmt_italic_btn.toggled.connect(
            lambda checked: self._apply_format_to_selection('italic', checked)
        )
        layout.addWidget(self.fmt_italic_btn)

        layout.addWidget(QLabel("Size:"))
        self.fmt_size_spin = QSpinBox()
        self.fmt_size_spin.setRange(6, 72)
        self.fmt_size_spin.setValue(10)
        self.fmt_size_spin.setFixedWidth(55)
        self.fmt_size_spin.setToolTip("Font size (pt)")
        self.fmt_size_spin.valueChanged.connect(
            lambda v: self._apply_format_to_selection('size', v)
        )
        layout.addWidget(self.fmt_size_spin)

        sep = QFrame()
        sep.setFrameShape(QFrame.VLine)
        sep.setFrameShadow(QFrame.Sunken)
        layout.addWidget(sep)

        self.fmt_fg_btn = QPushButton()
        self.fmt_fg_btn.setFixedSize(28, 26)
        self.fmt_fg_btn.setToolTip("Text Color")
        self.fmt_fg_btn.clicked.connect(self._pick_fg_color)
        layout.addWidget(self.fmt_fg_btn)

        self.fmt_bg_btn = QPushButton()
        self.fmt_bg_btn.setFixedSize(28, 26)
        self.fmt_bg_btn.setToolTip("Cell Background Color")
        self.fmt_bg_btn.clicked.connect(self._pick_bg_color)
        layout.addWidget(self.fmt_bg_btn)

        layout.addStretch()
        self._update_fmt_button_icons()
        return bar

    def _make_color_icon(self, label: str, color: QColor) -> QIcon:
        px = QPixmap(18, 18)
        px.fill(Qt.transparent)
        p = QPainter(px)
        font = QFont()
        font.setBold(True)
        font.setPointSize(8)
        p.setFont(font)
        p.setPen(Qt.black)
        p.drawText(QRect(0, 0, 18, 13), Qt.AlignCenter, label)
        p.fillRect(QRect(1, 14, 16, 3), color)
        p.end()
        return QIcon(px)

    def _update_fmt_button_icons(self):
        self.fmt_fg_btn.setIcon(self._make_color_icon("A", self._fmt_fg_color))
        self.fmt_fg_btn.setIconSize(QSize(18, 18))

        px = QPixmap(18, 18)
        px.fill(self._fmt_bg_color)
        p = QPainter(px)
        p.setPen(QColor("#888888"))
        p.drawRect(0, 0, 17, 17)
        p.end()
        self.fmt_bg_btn.setIcon(QIcon(px))
        self.fmt_bg_btn.setIconSize(QSize(18, 18))

    def _pick_fg_color(self):
        color = QColorDialog.getColor(self._fmt_fg_color, self, "Text Color")
        if color.isValid():
            self._fmt_fg_color = color
            self._update_fmt_button_icons()
            self._apply_format_to_selection('fg', color.name())

    def _pick_bg_color(self):
        color = QColorDialog.getColor(self._fmt_bg_color, self, "Background Color")
        if color.isValid():
            self._fmt_bg_color = color
            self._update_fmt_button_icons()
            self._apply_format_to_selection('bg', color.name())

    def _apply_format_to_selection(self, key: str, value):
        if self._source_model is None:
            return
        sel_model = self.table_view.selectionModel()
        if sel_model is None:
            return
        proxy_indices = sel_model.selectedIndexes()
        if not proxy_indices:
            return

        sheet_name = self._current_sheet_name() or '__csv__'
        sheet_fmt = self._all_sheet_formats.setdefault(sheet_name, {})

        try:
            self._source_model.itemChanged.disconnect(self._on_item_edited)
        except TypeError:
            pass

        for proxy_idx in proxy_indices:
            src_idx = self._proxy_model.mapToSource(proxy_idx)
            row, col = src_idx.row(), src_idx.column()
            item = self._source_model.item(row, col)
            if item is None:
                continue
            fmt = sheet_fmt.setdefault((row, col), {})
            explicit: set = set(item.data(_FMT_KEYS_ROLE) or set())
            if key == 'bold':
                font = item.font()
                font.setBold(value)
                item.setFont(font)
                fmt['bold'] = value
                explicit.add('bold')
            elif key == 'italic':
                font = item.font()
                font.setItalic(value)
                item.setFont(font)
                fmt['italic'] = value
                explicit.add('italic')
            elif key == 'size':
                font = item.font()
                font.setPointSize(value)
                item.setFont(font)
                fmt['size'] = value
                explicit.add('size')
            elif key == 'fg':
                item.setForeground(QBrush(QColor(value)))
                fmt['fg'] = value
                explicit.add('fg')
            elif key == 'bg':
                item.setBackground(QBrush(QColor(value)))
                fmt['bg'] = value
                explicit.add('bg')
            item.setData(explicit, _FMT_KEYS_ROLE)

            user_cell = self._user_changes.setdefault(sheet_name, {}).setdefault((row, col), {})
            user_cell[key] = value

        self._source_model.itemChanged.connect(self._on_item_edited)

    def _apply_format_to_item(self, item: QStandardItem, fmt: dict):
        explicit: set = set(item.data(_FMT_KEYS_ROLE) or set())
        if 'bold' in fmt or 'italic' in fmt or 'size' in fmt:
            font = item.font()
            if 'bold' in fmt:
                font.setBold(fmt['bold'])
                explicit.add('bold')
            if 'italic' in fmt:
                font.setItalic(fmt['italic'])
                explicit.add('italic')
            if 'size' in fmt:
                font.setPointSizeF(float(fmt['size']))
                explicit.add('size')
            item.setFont(font)
        if 'fg' in fmt:
            item.setForeground(QBrush(QColor(fmt['fg'])))
            explicit.add('fg')
        if 'bg' in fmt:
            item.setBackground(QBrush(QColor(fmt['bg'])))
            explicit.add('bg')
        item.setData(explicit, _FMT_KEYS_ROLE)

    def _fmt_from_item(self, item: QStandardItem) -> dict:
        d = {}
        explicit = item.data(_FMT_KEYS_ROLE)
        if not explicit:
            return d
        if 'bold' in explicit or 'italic' in explicit or 'size' in explicit:
            font = item.font()
            if 'bold' in explicit:
                d['bold'] = font.bold()
            if 'italic' in explicit:
                d['italic'] = font.italic()
            if 'size' in explicit and font.pointSize() > 0:
                d['size'] = font.pointSize()
        if 'fg' in explicit:
            fg = item.data(Qt.ForegroundRole)
            if fg is not None:
                c = fg.color() if isinstance(fg, QBrush) else QColor(fg)
                if c.isValid():
                    d['fg'] = c.name()
        if 'bg' in explicit:
            bg = item.data(Qt.BackgroundRole)
            if bg is not None:
                c = bg.color() if isinstance(bg, QBrush) else QColor(bg)
                if c.isValid() and c.alpha() > 0:
                    d['bg'] = c.name()
        return d

    def _sync_sheet_fmt_from_model(self):
        if self._source_model is None:
            return
        sheet_name = self._current_sheet_name() or '__csv__'
        sheet_fmt: dict = {}
        for row in range(self._source_model.rowCount()):
            for col in range(self._source_model.columnCount()):
                item = self._source_model.item(row, col)
                if item:
                    fmt = self._fmt_from_item(item)
                    if fmt:
                        sheet_fmt[(row, col)] = fmt
        self._all_sheet_formats[sheet_name] = sheet_fmt

    def _on_current_cell_changed(self, current, _previous):
        if not current.isValid() or self._source_model is None:
            return
        src_idx = self._proxy_model.mapToSource(current)
        row, col = src_idx.row(), src_idx.column()

        sheet_name = self._current_sheet_name() or '__csv__'
        fmt = self._all_sheet_formats.get(sheet_name, {}).get((row, col), {})

        self.fmt_bold_btn.blockSignals(True)
        self.fmt_bold_btn.setChecked(fmt.get('bold', False))
        self.fmt_bold_btn.blockSignals(False)

        self.fmt_italic_btn.blockSignals(True)
        self.fmt_italic_btn.setChecked(fmt.get('italic', False))
        self.fmt_italic_btn.blockSignals(False)

        self.fmt_size_spin.blockSignals(True)
        item = self._source_model.item(row, col)
        if item:
            size = item.font().pointSize()
            if size <= 0:
                size = QApplication.font().pointSize()
            if size > 0:
                self.fmt_size_spin.setValue(size)
        self.fmt_size_spin.blockSignals(False)

        self._fmt_fg_color = QColor(fmt['fg']) if 'fg' in fmt else QColor(Qt.black)
        self._fmt_bg_color = QColor(fmt['bg']) if 'bg' in fmt else QColor(Qt.white)
        self._update_fmt_button_icons()

    # ------------------------------------------------------------------
    # Undo
    # ------------------------------------------------------------------

    def _undo(self):
        if not self._undo_stack or self._source_model is None:
            return
        row, col, old_text = self._undo_stack.pop()
        item = self._source_model.item(row, col)
        if item is None:
            return

        self._set_df_cell(self.df, row, col, old_text)
        sheet_name = self._current_sheet_name()
        if sheet_name and sheet_name in self._sheet_cache:
            self._set_df_cell(self._sheet_cache[sheet_name], row, col, old_text)

        self._cell_shadow[(row, col)] = old_text

        try:
            self._source_model.itemChanged.disconnect(self._on_item_edited)
        except TypeError:
            pass
        item.setText(old_text)
        self._source_model.itemChanged.connect(self._on_item_edited)

    # ------------------------------------------------------------------
    # xlsx formatting I/O
    # ------------------------------------------------------------------

    def _read_xlsx_formatting(self, file_path: str, sheet_name: str) -> dict:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                return {}
            ws = wb[sheet_name]
            fmt = {}
            for row_idx, row in enumerate(ws.iter_rows(min_row=2)):
                for col_idx, cell in enumerate(row):
                    d = {}
                    f = cell.font
                    if f:
                        if f.bold:
                            d['bold'] = True
                        if f.italic:
                            d['italic'] = True
                        if f.size:
                            d['size'] = float(f.size)
                        if f.color and f.color.type == 'rgb':
                            rgb = f.color.rgb
                            if rgb and rgb.upper() not in ('00000000', 'FF000000'):
                                d['fg'] = '#' + rgb[2:]
                    fill = cell.fill
                    if fill and fill.fill_type == 'solid':
                        fc = fill.fgColor
                        if fc and fc.type == 'rgb':
                            rgb = fc.rgb
                            if rgb and rgb.upper() not in ('00000000', 'FFFFFFFF'):
                                d['bg'] = '#' + rgb[2:]
                    if d:
                        fmt[(row_idx, col_idx)] = d
            return fmt
        except Exception:
            traceback.print_exc()
            return {}

    def _apply_openpyxl_fmt(self, cell, fmt: dict):
        from openpyxl.styles import Font, PatternFill
        if any(k in fmt for k in ('bold', 'italic', 'size', 'fg')):
            kw = {}
            if 'bold'   in fmt: kw['bold']   = fmt['bold']
            if 'italic' in fmt: kw['italic'] = fmt['italic']
            if 'size'   in fmt: kw['size']   = fmt['size']
            if 'fg'     in fmt: kw['color']  = 'FF' + fmt['fg'][1:]
            cell.font = Font(**kw)
        if 'bg' in fmt:
            cell.fill = PatternFill(fill_type='solid', fgColor='FF' + fmt['bg'][1:])

    def _apply_fmt_merged(self, cell, user_fmt: dict):
        from openpyxl.styles import Font, PatternFill
        from copy import copy
        if any(k in user_fmt for k in ('bold', 'italic', 'size', 'fg')):
            f = copy(cell.font)
            kw = dict(
                name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                underline=f.underline, strike=f.strike, color=f.color,
                vertAlign=f.vertAlign, charset=f.charset, family=f.family,
                scheme=f.scheme,
            )
            if 'bold'   in user_fmt: kw['bold']   = user_fmt['bold']
            if 'italic' in user_fmt: kw['italic'] = user_fmt['italic']
            if 'size'   in user_fmt: kw['size']   = user_fmt['size']
            if 'fg'     in user_fmt: kw['color']  = 'FF' + user_fmt['fg'][1:]
            cell.font = Font(**kw)
        if 'bg' in user_fmt:
            cell.fill = PatternFill(fill_type='solid',
                                    fgColor='FF' + user_fmt['bg'][1:])

    # ------------------------------------------------------------------
    # Global search
    # ------------------------------------------------------------------

    def _on_global_search_text_changed(self, text: str):
        self.search_clear_btn.setVisible(bool(text))
        self._search_timer.start()

    def _apply_global_search(self):
        if self._proxy_model:
            self._proxy_model.set_global_filter(self.global_search_input.text())
            self._update_status()

    def _clear_global_search(self):
        self.global_search_input.clear()

    # ------------------------------------------------------------------
    # DataFrame helpers
    # ------------------------------------------------------------------

    def _set_df_cell(self, df, row: int, col: int, text: str):
        if df is None or row >= len(df) or col >= len(df.columns):
            return
        col_name = df.columns[col]
        dtype = df[col_name].dtype
        try:
            if pd.api.types.is_integer_dtype(dtype):
                value = int(text) if text.strip() else None
            elif pd.api.types.is_float_dtype(dtype):
                value = float(text) if text.strip() else None
            else:
                value = text
        except (ValueError, TypeError):
            value = text
        try:
            df.iat[row, col] = value
        except (TypeError, ValueError):
            df[col_name] = df[col_name].astype(object)
            df.iat[row, col] = value

    def _on_item_edited(self, item):
        row, col = item.row(), item.column()
        text = item.text()

        old_text = self._cell_shadow.get((row, col), "")
        if old_text == text:
            return

        self._undo_stack.append((row, col, old_text))
        self._cell_shadow[(row, col)] = text

        if self.df is None:
            return

        self._set_df_cell(self.df, row, col, text)

        if self._excel_sheets:
            idx = self.sheet_tab_bar.currentIndex()
            if 0 <= idx < len(self._excel_sheets):
                sheet_name = self._excel_sheets[idx]
                if sheet_name in self._sheet_cache:
                    self._set_df_cell(self._sheet_cache[sheet_name], row, col, text)

    # ------------------------------------------------------------------
    # Context menu
    # ------------------------------------------------------------------

    def _show_context_menu(self, pos: QPoint):
        if self._source_model is None:
            return

        selected = self.table_view.selectionModel().selectedIndexes()
        if not selected:
            return

        proxy = self._proxy_model
        source_rows = sorted({proxy.mapToSource(i).row() for i in selected})
        source_cols = sorted({proxy.mapToSource(i).column() for i in selected})

        insert_row = source_rows[-1] + 1
        insert_col = source_cols[-1] + 1

        menu = QMenu(self)

        act_copy  = menu.addAction("Copy\tCtrl+C")
        act_cut   = menu.addAction("Cut\tCtrl+X")
        act_paste = menu.addAction("Paste\tCtrl+V")

        menu.addSeparator()

        row_label = f"Delete {len(source_rows)} Row(s)" if len(source_rows) > 1 else "Delete Row"
        act_del_rows = menu.addAction(row_label)

        act_add_row = menu.addAction("Add Row Below")

        menu.addSeparator()

        col_label = f"Delete {len(source_cols)} Column(s)" if len(source_cols) > 1 else "Delete Column"
        act_del_cols = menu.addAction(col_label)

        act_add_col = menu.addAction("Add Column Right…")

        action = menu.exec_(self.table_view.viewport().mapToGlobal(pos))

        if action == act_copy:
            self._copy_selection()
        elif action == act_cut:
            self._cut_selection()
        elif action == act_paste:
            self._paste_selection()
        elif action == act_del_rows:
            self._delete_rows(source_rows)
        elif action == act_add_row:
            self._add_row(insert_row)
        elif action == act_del_cols:
            self._delete_columns(source_cols)
        elif action == act_add_col:
            self._add_column(insert_col)

    # ------------------------------------------------------------------
    # Marching-ants helpers
    # ------------------------------------------------------------------

    def eventFilter(self, obj, event):
        if obj is self.table_view.viewport() and event.type() == QEvent.Resize:
            self._march_overlay.resize(obj.size())
        return super().eventFilter(obj, event)

    def _march_tick(self):
        self._march_offset = (self._march_offset + 1.5) % 8.0
        self._update_march_overlay()

    def _update_march_overlay(self):
        if self._clipboard_region is None or self._proxy_model is None:
            self._march_overlay.clear_rect()
            return
        rows = self._clipboard_region['rows']
        cols = self._clipboard_region['cols']
        tl = self._proxy_model.mapFromSource(self._source_model.index(rows[0],  cols[0]))
        br = self._proxy_model.mapFromSource(self._source_model.index(rows[-1], cols[-1]))
        if not tl.isValid() or not br.isValid():
            self._march_overlay.clear_rect()
            return
        rect = self.table_view.visualRect(tl).united(self.table_view.visualRect(br))
        if rect.isValid():
            self._march_overlay.set_rect(rect, self._march_offset)
        else:
            self._march_overlay.clear_rect()

    def _cancel_copy_cut(self):
        self._clipboard_region = None
        self._march_timer.stop()
        self._march_overlay.clear_rect()

    # ------------------------------------------------------------------
    # Cut / Copy / Paste
    # ------------------------------------------------------------------

    def _capture_region(self, mode: str):
        if self._source_model is None:
            return False
        indexes = self.table_view.selectionModel().selectedIndexes()
        if not indexes:
            return False
        src = [self._proxy_model.mapToSource(i) for i in indexes]
        rows = sorted({i.row() for i in src})
        cols = sorted({i.column() for i in src})
        data, fmt = [], {}
        for dr, r in enumerate(rows):
            row_data = []
            for dc, c in enumerate(cols):
                item = self._source_model.item(r, c)
                val  = item.text() if item else ""
                row_data.append(val)
                if item:
                    cell_fmt = self._fmt_from_item(item)
                    if cell_fmt:
                        fmt[(dr, dc)] = cell_fmt
            data.append(row_data)
        self._clipboard_region = {'mode': mode, 'rows': rows, 'cols': cols,
                                  'data': data, 'fmt': fmt}
        QApplication.clipboard().setText(
            "\n".join("\t".join(row) for row in data)
        )
        self._march_offset = 0.0
        self._march_timer.start()
        self._update_march_overlay()
        return True

    def _copy_selection(self):
        self._capture_region('copy')

    def _cut_selection(self):
        self._capture_region('cut')

    def _paste_selection(self):
        if self._source_model is None:
            return
        indexes = self.table_view.selectionModel().selectedIndexes()
        if not indexes:
            return
        src_indices = [self._proxy_model.mapToSource(i) for i in indexes]
        start_row = min(i.row() for i in src_indices)
        start_col = min(i.column() for i in src_indices)
        sheet_name = self._current_sheet_name()

        if self._clipboard_region is not None:
            region = self._clipboard_region
            try:
                self._source_model.itemChanged.disconnect(self._on_item_edited)
            except TypeError:
                pass
            for dr, row_data in enumerate(region['data']):
                r = start_row + dr
                if r >= self._source_model.rowCount():
                    break
                for dc, val in enumerate(row_data):
                    c = start_col + dc
                    if c >= self._source_model.columnCount():
                        break
                    item = self._source_model.item(r, c)
                    if item is None:
                        item = QStandardItem()
                        item.setEditable(True)
                        self._source_model.setItem(r, c, item)
                    item.setText(val)
                    cell_fmt = region['fmt'].get((dr, dc), {})
                    if cell_fmt:
                        self._apply_format_to_item(item, cell_fmt)
                        uc = self._user_changes.setdefault(
                            sheet_name or '__csv__', {}).setdefault((r, c), {})
                        uc.update(cell_fmt)
                    self._set_df_cell(self.df, r, c, val)
                    if sheet_name and sheet_name in self._sheet_cache:
                        self._set_df_cell(self._sheet_cache[sheet_name], r, c, val)
            self._source_model.itemChanged.connect(self._on_item_edited)

            if region['mode'] == 'cut':
                try:
                    self._source_model.itemChanged.disconnect(self._on_item_edited)
                except TypeError:
                    pass
                for r in region['rows']:
                    for c in region['cols']:
                        item = self._source_model.item(r, c)
                        if item:
                            item.setText("")
                        if self.df is not None and r < len(self.df) and c < len(self.df.columns):
                            self.df.iat[r, c] = ""
                        if sheet_name and sheet_name in self._sheet_cache:
                            df_c = self._sheet_cache[sheet_name]
                            if r < len(df_c) and c < len(df_c.columns):
                                df_c.iat[r, c] = ""
                self._source_model.itemChanged.connect(self._on_item_edited)

            self._cancel_copy_cut()
        else:
            text = QApplication.clipboard().text()
            if not text:
                return
            try:
                self._source_model.itemChanged.disconnect(self._on_item_edited)
            except TypeError:
                pass
            for dr, line in enumerate(text.split("\n")):
                r = start_row + dr
                if r >= self._source_model.rowCount():
                    break
                for dc, val in enumerate(line.split("\t")):
                    c = start_col + dc
                    if c >= self._source_model.columnCount():
                        break
                    item = self._source_model.item(r, c)
                    if item is None:
                        item = QStandardItem()
                        item.setEditable(True)
                        self._source_model.setItem(r, c, item)
                    item.setText(val)
                    self._set_df_cell(self.df, r, c, val)
                    if sheet_name and sheet_name in self._sheet_cache:
                        self._set_df_cell(self._sheet_cache[sheet_name], r, c, val)
            self._source_model.itemChanged.connect(self._on_item_edited)

    # ------------------------------------------------------------------
    # Structural edits (add / delete rows & columns)
    # ------------------------------------------------------------------

    def _delete_rows(self, source_rows: list):
        try:
            sheet_name = self._current_sheet_name()
            rows_desc = sorted(source_rows, reverse=True)

            self._structural_ops.append((sheet_name, 'delete_rows', rows_desc))

            if self.df is not None:
                self.df = self.df.drop(self.df.index[rows_desc]).reset_index(drop=True)
            if sheet_name and sheet_name in self._sheet_cache:
                df_c = self._sheet_cache[sheet_name]
                self._sheet_cache[sheet_name] = df_c.drop(df_c.index[rows_desc]).reset_index(drop=True)

            for fmt_dict in (self._all_sheet_formats.get(sheet_name, {}),
                             self._user_changes.get(sheet_name, {})):
                self._remap_rows(fmt_dict, source_rows)

            for row in rows_desc:
                self._source_model.removeRow(row)

            self._update_status()
        except Exception:
            traceback.print_exc()

    def _delete_columns(self, source_cols: list):
        try:
            sheet_name = self._current_sheet_name()
            cols_desc = sorted(source_cols, reverse=True)

            self._structural_ops.append((sheet_name, 'delete_cols', cols_desc))

            if self.df is not None:
                self.df = self.df.drop(self.df.columns[cols_desc], axis=1)
            if sheet_name and sheet_name in self._sheet_cache:
                df_c = self._sheet_cache[sheet_name]
                self._sheet_cache[sheet_name] = df_c.drop(df_c.columns[cols_desc], axis=1)

            for fmt_dict in (self._all_sheet_formats.get(sheet_name, {}),
                             self._user_changes.get(sheet_name, {})):
                self._remap_cols(fmt_dict, source_cols)

            for col in cols_desc:
                self._source_model.removeColumn(col)

            self._update_status()
        except Exception:
            traceback.print_exc()

    def _add_row(self, insert_at: int):
        try:
            sheet_name = self._current_sheet_name()

            self._structural_ops.append((sheet_name, 'insert_row', insert_at))

            if self.df is not None:
                blank = pd.DataFrame([[""] * len(self.df.columns)], columns=self.df.columns)
                self.df = pd.concat(
                    [self.df.iloc[:insert_at], blank, self.df.iloc[insert_at:]],
                    ignore_index=True
                )
            if sheet_name and sheet_name in self._sheet_cache:
                df_c = self._sheet_cache[sheet_name]
                blank = pd.DataFrame([[""] * len(df_c.columns)], columns=df_c.columns)
                self._sheet_cache[sheet_name] = pd.concat(
                    [df_c.iloc[:insert_at], blank, df_c.iloc[insert_at:]],
                    ignore_index=True
                )

            for fmt_dict in (self._all_sheet_formats.get(sheet_name, {}),
                             self._user_changes.get(sheet_name, {})):
                self._shift_rows(fmt_dict, insert_at)

            col_count = self._source_model.columnCount()
            items = [QStandardItem("") for _ in range(col_count)]
            for it in items:
                it.setEditable(True)
            self._source_model.insertRow(insert_at, items)

            self._update_status()
        except Exception:
            traceback.print_exc()

    def _add_column(self, insert_at: int):
        try:
            name, ok = QInputDialog.getText(self, "Add Column", "Column name:")
            if not ok or not name.strip():
                return
            name = name.strip()

            sheet_name = self._current_sheet_name()

            self._structural_ops.append((sheet_name, 'insert_col', insert_at))

            if self.df is not None:
                self.df.insert(insert_at, name, "")
            if sheet_name and sheet_name in self._sheet_cache:
                df_c = self._sheet_cache[sheet_name]
                if df_c is not self.df:
                    df_c.insert(insert_at, name, "")

            for fmt_dict in (self._all_sheet_formats.get(sheet_name, {}),
                             self._user_changes.get(sheet_name, {})):
                self._shift_cols(fmt_dict, insert_at)

            self._source_model.insertColumn(insert_at)
            self._source_model.setHorizontalHeaderItem(insert_at, QStandardItem(name))
            self._source_model.setHeaderData(
                insert_at, Qt.Horizontal,
                int(Qt.AlignLeft | Qt.AlignVCenter),
                Qt.TextAlignmentRole
            )
            for row in range(self._source_model.rowCount()):
                item = QStandardItem("")
                item.setEditable(True)
                self._source_model.setItem(row, insert_at, item)

            self._update_status()
        except Exception:
            traceback.print_exc()

    # helpers ---------------------------------------------------------------

    def _current_sheet_name(self) -> str:
        if self._excel_sheets:
            idx = self.sheet_tab_bar.currentIndex()
            if 0 <= idx < len(self._excel_sheets):
                return self._excel_sheets[idx]
        return ""

    @staticmethod
    def _remap_rows(fmt_dict: dict, deleted_rows: list):
        deleted = set(deleted_rows)
        new_dict = {}
        for (r, c), v in fmt_dict.items():
            if r in deleted:
                continue
            shift = sum(1 for d in deleted if d < r)
            new_dict[(r - shift, c)] = v
        fmt_dict.clear()
        fmt_dict.update(new_dict)

    @staticmethod
    def _remap_cols(fmt_dict: dict, deleted_cols: list):
        deleted = set(deleted_cols)
        new_dict = {}
        for (r, c), v in fmt_dict.items():
            if c in deleted:
                continue
            shift = sum(1 for d in deleted if d < c)
            new_dict[(r, c - shift)] = v
        fmt_dict.clear()
        fmt_dict.update(new_dict)

    @staticmethod
    def _shift_rows(fmt_dict: dict, insert_at: int):
        new_dict = {}
        for (r, c), v in fmt_dict.items():
            new_dict[(r + 1 if r >= insert_at else r, c)] = v
        fmt_dict.clear()
        fmt_dict.update(new_dict)

    @staticmethod
    def _shift_cols(fmt_dict: dict, insert_at: int):
        new_dict = {}
        for (r, c), v in fmt_dict.items():
            new_dict[(r, c + 1 if c >= insert_at else c)] = v
        fmt_dict.clear()
        fmt_dict.update(new_dict)

    # ------------------------------------------------------------------
    # Sort / Filter
    # ------------------------------------------------------------------

    def _on_sort_requested(self, col: int, order: int):
        if self._proxy_model:
            self._proxy_model.sort(col, order)

    def _show_filter_popup(self, col_index: int):
        if self._source_model is None:
            return

        all_values: set = set()
        for row in range(self._source_model.rowCount()):
            idx = self._source_model.index(row, col_index)
            all_values.add(self._source_model.data(idx, Qt.DisplayRole) or "")

        numeric        = _column_is_numeric(list(all_values))
        current_filter = self._proxy_model.get_column_filter(col_index)

        sec_x      = self._header.sectionViewportPosition(col_index)
        global_pos = self._header.mapToGlobal(QPoint(sec_x, self._header.height()))

        self._filter_popup = FilterPopup(
            col_index, list(all_values),
            current_filter, is_numeric=numeric
        )
        self._filter_popup.filter_changed.connect(self._on_filter_changed)
        self._filter_popup.move(global_pos)
        self._filter_popup.show()

    def _on_filter_changed(self, col: int, spec):
        self._proxy_model.set_column_filter(col, spec)
        self._header.mark_filter_active(col, spec is not None)
        self._update_status()

    # ------------------------------------------------------------------
    # Sheet tabs
    # ------------------------------------------------------------------

    def _setup_sheet_tabs(self, sheets: list):
        self.sheet_tab_bar.blockSignals(True)
        while self.sheet_tab_bar.count():
            self.sheet_tab_bar.removeTab(0)
        for name in sheets:
            self.sheet_tab_bar.addTab(name)
        self.sheet_tab_bar.setCurrentIndex(0)
        self.sheet_tab_bar.setVisible(len(sheets) > 1)
        self.sheet_tab_bar.blockSignals(False)

    def _on_sheet_tab_changed(self, index: int):
        if not self._excel_sheets or index < 0:
            return
        self._sync_sheet_fmt_from_model()
        sheet_name = self._excel_sheets[index]
        if sheet_name not in self._sheet_cache:
            try:
                df = pd.read_excel(self.current_file_path, sheet_name=sheet_name)
                self._sheet_cache[sheet_name] = df
            except Exception as e:
                traceback.print_exc()
                QMessageBox.critical(self, "Error", f"Could not load sheet '{sheet_name}':\n{e}")
                return
        self._load_dataframe(self._sheet_cache[sheet_name],
                             self.current_file_path, sheet_name=sheet_name)

    def _on_clear_filter(self, col: int):
        self._proxy_model.set_column_filter(col, None)
        self._header.mark_filter_active(col, False)
        self._update_status()

    # ------------------------------------------------------------------
    # File loading
    # ------------------------------------------------------------------

    def show_open_dialog(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open File", "",
            "All Supported (*.xlsx *.xls *.csv);;"
            "Excel Files (*.xlsx *.xls);;"
            "CSV Files (*.csv);;"
            "All Files (*)"
        )
        if file_path:
            self.open_file(file_path)

    def open_file(self, file_path: str):
        ext = os.path.splitext(file_path)[1].lower()
        try:
            if ext in ('.xlsx', '.xls'):
                xf = pd.ExcelFile(file_path)
                sheets = xf.sheet_names
                self._excel_sheets      = sheets
                self._sheet_cache       = {}
                self._all_sheet_formats = {}
                self._user_changes      = {}
                self._structural_ops    = []
                df = xf.parse(sheets[0])
                self._sheet_cache[sheets[0]] = df
                self._setup_sheet_tabs(sheets)
                self._load_dataframe(df, file_path, sheet_name=sheets[0])
            elif ext == '.csv':
                self._excel_sheets      = []
                self._sheet_cache       = {}
                self._all_sheet_formats = {}
                self._user_changes      = {}
                self._structural_ops    = []
                self._setup_sheet_tabs([])
                df = self._read_csv_with_auto_encoding(file_path)
                self._load_dataframe(df, file_path)
            else:
                QMessageBox.warning(
                    self, "Unsupported File",
                    f"Unsupported file type: {ext}\n\nSupported: .xlsx, .xls, .csv"
                )
                return
        except Exception as e:
            traceback.print_exc()
            QMessageBox.critical(
                self, "Error Opening File",
                f"Could not open file:\n{file_path}\n\n{e}"
            )

    def _read_csv_with_auto_encoding(self, file_path: str) -> pd.DataFrame:
        for enc in CSV_ENCODINGS:
            try:
                return pd.read_csv(file_path, encoding=enc)
            except UnicodeDecodeError:
                continue
        raise ValueError(
            "Could not decode the CSV file.\n"
            f"Tried encodings: {', '.join(CSV_ENCODINGS)}"
        )

    def _load_blank_sheet(self, rows: int = 10, cols: int = 5):
        col_names = [chr(ord('A') + i) for i in range(cols)]
        df = pd.DataFrame("", index=range(rows), columns=col_names)
        self._excel_sheets      = []
        self._sheet_cache       = {}
        self._all_sheet_formats = {}
        self._user_changes      = {}
        self._structural_ops    = []
        self._setup_sheet_tabs([])
        self._load_dataframe(df, "")

    def _load_dataframe(self, df: pd.DataFrame, file_path: str, sheet_name: str = ""):
        self.df                = df
        self.current_file_path = file_path

        source = QStandardItemModel(len(df), len(df.columns))
        source.setHorizontalHeaderLabels(df.columns.astype(str).tolist())

        left_vcenter = int(Qt.AlignLeft | Qt.AlignVCenter)
        for col in range(len(df.columns)):
            source.setHeaderData(col, Qt.Horizontal, left_vcenter, Qt.TextAlignmentRole)

        for row_idx, (_, row_data) in enumerate(df.iterrows()):
            for col_idx, value in enumerate(row_data):
                display = "" if pd.isna(value) else str(value)
                item = QStandardItem(display)
                item.setEditable(True)
                source.setItem(row_idx, col_idx, item)

        self._source_model = source
        self._proxy_model  = MultiColumnFilterProxyModel()
        self._proxy_model.setSourceModel(source)

        self.table_view.setModel(self._proxy_model)
        self.table_view.selectionModel().currentChanged.connect(self._on_current_cell_changed)

        self.table_view.resizeColumnsToContents()
        self._ensure_header_column_widths()

        self._header.reset_state()

        self._undo_stack = []
        self._cell_shadow = {}
        for r in range(self._source_model.rowCount()):
            for c in range(self._source_model.columnCount()):
                itm = self._source_model.item(r, c)
                if itm:
                    self._cell_shadow[(r, c)] = itm.text()

        self.table_view.setEditTriggers(
            QTableView.DoubleClicked | QTableView.AnyKeyPressed
        )
        self._source_model.itemChanged.connect(self._on_item_edited)

        ext = os.path.splitext(file_path)[1].lower()
        if ext in ('.xlsx', '.xls') and sheet_name:
            sheet_fmt = self._all_sheet_formats.setdefault(sheet_name, {})
            if not sheet_fmt:
                raw = self._read_xlsx_formatting(file_path, sheet_name)
                sheet_fmt.update(raw)
            for (r, c), fmt in list(sheet_fmt.items()):
                if r < self._source_model.rowCount() and c < self._source_model.columnCount():
                    item = self._source_model.item(r, c)
                    if item and fmt:
                        self._apply_format_to_item(item, fmt)

        self._search_timer.stop()
        self.global_search_input.blockSignals(True)
        self.global_search_input.clear()
        self.global_search_input.blockSignals(False)
        self.search_clear_btn.setVisible(False)

        file_name = os.path.basename(file_path)
        title = f"Table Viewer — {file_name}" if file_name else "Table Viewer"
        if sheet_name:
            title += f"  [{sheet_name}]"
        self.setWindowTitle(title)
        self._update_status()

    def _ensure_header_column_widths(self):
        fm        = self._header.fontMetrics()
        reserve   = self._header._ICON_RESERVE
        h_padding = 16

        for col in range(self._source_model.columnCount()):
            label = self._source_model.headerData(col, Qt.Horizontal, Qt.DisplayRole) or ""
            min_w = fm.horizontalAdvance(str(label)) + reserve + h_padding
            if self.table_view.columnWidth(col) < min_w:
                self.table_view.setColumnWidth(col, min_w)

    def _update_status(self):
        if self.df is None:
            return
        file_name = os.path.basename(self.current_file_path) if self.current_file_path else ""
        total   = len(self.df)
        visible = self._proxy_model.rowCount() if self._proxy_model else total
        cols    = len(self.df.columns)
        label   = f"{file_name}   |   " if file_name else ""
        if visible == total:
            self.status_label.setText(f"{label}{total:,} rows   |   {cols} columns")
        else:
            self.status_label.setText(
                f"{label}{visible:,} / {total:,} rows (filtered)   |   {cols} columns"
            )

    # ------------------------------------------------------------------
    # File saving
    # ------------------------------------------------------------------

    def save_as_excel(self):
        if self.df is None:
            return
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save as Excel", self._default_save_name('.xlsx'),
            "Excel Files (*.xlsx);;All Files (*)"
        )
        if not file_path:
            return
        try:
            import openpyxl
            src_ext = os.path.splitext(self.current_file_path or '')[1].lower()

            if src_ext in ('.xlsx', '.xls') and self.current_file_path:
                wb = openpyxl.load_workbook(self.current_file_path)
                sheets = self._excel_sheets if self._excel_sheets else (
                    [wb.sheetnames[0]] if wb.sheetnames else ['Sheet1']
                )

                for sn, op, arg in self._structural_ops:
                    target = sn if sn else (wb.sheetnames[0] if wb.sheetnames else None)
                    if target is None or target not in wb.sheetnames:
                        continue
                    ws = wb[target]
                    if op == 'insert_row':
                        ws.insert_rows(arg + 2)
                    elif op == 'delete_rows':
                        for r in arg:
                            ws.delete_rows(r + 2)
                    elif op == 'insert_col':
                        ws.insert_cols(arg + 1)
                    elif op == 'delete_cols':
                        for c in arg:
                            ws.delete_cols(c + 1)

                for sheet_name in sheets:
                    ws = (wb[sheet_name] if sheet_name in wb.sheetnames
                          else wb.create_sheet(title=sheet_name))
                    df = self._sheet_cache.get(sheet_name) if sheet_name else self.df
                    if df is None:
                        df = self.df
                    if df is None:
                        continue

                    for ci, col_name in enumerate(df.columns, 1):
                        ws.cell(row=1, column=ci).value = str(col_name)
                    for ri, (_, row_data) in enumerate(df.iterrows(), 2):
                        for ci, val in enumerate(row_data):
                            ws.cell(row=ri, column=ci + 1).value = (
                                None if pd.isna(val) else val
                            )

                    user_fmt = self._user_changes.get(sheet_name, {})
                    for (r, c), fmt in user_fmt.items():
                        self._apply_fmt_merged(ws.cell(row=r + 2, column=c + 1), fmt)

                wb.save(file_path)
                self.current_file_path = file_path
                self._structural_ops = []
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                sheet_name = self._excel_sheets[0] if self._excel_sheets else '__csv__'
                df = self.df

                for ci, col_name in enumerate(df.columns, 1):
                    ws.cell(row=1, column=ci).value = str(col_name)
                for ri, (_, row_data) in enumerate(df.iterrows(), 2):
                    for ci, val in enumerate(row_data, 1):
                        ws.cell(row=ri, column=ci).value = (
                            None if pd.isna(val) else val
                        )

                user_fmt = self._user_changes.get(sheet_name, {})
                for (r, c), fmt in user_fmt.items():
                    self._apply_fmt_merged(ws.cell(row=r + 2, column=c + 1), fmt)

                wb.save(file_path)
                self.current_file_path = file_path
                self._structural_ops = []

            n = len(self._excel_sheets)
            self.statusBar().showMessage(
                f"Saved {n} sheets." if n > 1 else "Saved.", 3000
            )
        except Exception as e:
            traceback.print_exc()
            QMessageBox.critical(self, "Error Saving File", str(e))

    def save_as_csv(self):
        if self.df is None:
            return
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save as CSV", self._default_save_name('.csv'),
            "CSV Files (*.csv);;All Files (*)"
        )
        if file_path:
            try:
                self.df.to_csv(file_path, index=False, encoding='utf-8-sig')
                self.statusBar().showMessage("Saved successfully.", 3000)
            except Exception as e:
                traceback.print_exc()
                QMessageBox.critical(self, "Error Saving File", str(e))

    def _default_save_name(self, new_ext: str) -> str:
        if self.current_file_path:
            return os.path.splitext(self.current_file_path)[0] + new_ext
        return ""

    # ------------------------------------------------------------------
    # Drag & Drop
    # ------------------------------------------------------------------

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            url = event.mimeData().urls()[0].toLocalFile()
            if url.lower().endswith(SUPPORTED_EXTENSIONS):
                event.acceptProposedAction()

    def dropEvent(self, event):
        self.open_file(event.mimeData().urls()[0].toLocalFile())

    # ------------------------------------------------------------------
    # File association  (Windows Registry)
    # ------------------------------------------------------------------

    def register_file_associations(self):
        try:
            import winreg
        except ImportError:
            QMessageBox.warning(self, "Not Supported",
                                "File association registration is only supported on Windows.")
            return

        app_path = os.path.abspath(sys.argv[0])
        command  = (f'"{sys.executable}" "{app_path}" "%1"'
                    if app_path.endswith('.py')
                    else f'"{app_path}" "%1"')

        try:
            for ext in SUPPORTED_EXTENSIONS:
                prog_id = f"TableViewer.{ext[1:].upper()}"

                with winreg.CreateKey(
                    winreg.HKEY_CURRENT_USER,
                    rf"Software\Classes\{prog_id}\shell\open\command"
                ) as k:
                    winreg.SetValueEx(k, "", 0, winreg.REG_SZ, command)

                icon_src = app_path if not app_path.endswith('.py') else sys.executable
                with winreg.CreateKey(
                    winreg.HKEY_CURRENT_USER,
                    rf"Software\Classes\{prog_id}\DefaultIcon"
                ) as k:
                    winreg.SetValueEx(k, "", 0, winreg.REG_SZ, f'"{icon_src}",0')

                with winreg.CreateKey(
                    winreg.HKEY_CURRENT_USER,
                    rf"Software\Classes\{ext}\OpenWithProgids"
                ) as k:
                    winreg.SetValueEx(k, prog_id, 0, winreg.REG_NONE, b"")

            QMessageBox.information(
                self, "File Associations Registered",
                "Done! .xlsx, .xls and .csv files are now associated with Table Viewer.\n\n"
                "To set as default:\n"
                "  Right-click any .xlsx/.xls/.csv file\n"
                "  → Open with → Choose another app\n"
                "  → Table Viewer → Always use this app"
            )
        except Exception as e:
            traceback.print_exc()
            QMessageBox.critical(self, "Registration Failed",
                                 f"Could not register file associations:\n\n{e}")
